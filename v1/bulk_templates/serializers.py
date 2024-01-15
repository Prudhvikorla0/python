"""
Serializers for notifications=
"""

import openpyxl
from sentry_sdk import capture_exception

from django.utils.translation import gettext_lazy as _

from rest_framework import serializers

from base import session
from base import exceptions

from v1.supply_chains import models as sc_models

from common.drf_custom import fields

from v1.bulk_templates import models as bulk_models 
from . import constants

from v1.tenants import models as tenant_models


class BulkUploadSerializer(serializers.ModelSerializer):
    """
    Serializer to validate a template
    """
    id = fields.IdencodeField(read_only=True)
    status = serializers.IntegerField(read_only=True)
    created_on = serializers.DateTimeField(read_only=True)
    updated_on = serializers.DateTimeField(read_only=True)
    total_new_items = serializers.IntegerField(read_only=True)
    new_items_completed = serializers.IntegerField(read_only=True)
    new_items_failed = serializers.IntegerField(read_only=True)
    total_updations = serializers.IntegerField(read_only=True)
    updations_completed = serializers.IntegerField(read_only=True)
    updations_failed = serializers.IntegerField(read_only=True)
    file = serializers.FileField(write_only=True)
    file_name = serializers.CharField(required=False)
    type = serializers.IntegerField(write_only=True)
    supply_chain = fields.IdencodeField(
        write_only=True, related_model=sc_models.SupplyChain)
    validated_data = serializers.JSONField(read_only=True)
    data = serializers.JSONField(required=False)
    is_valid = serializers.BooleanField(read_only=True)

    class Meta:
        fields = (
            'id', 'file', 'file_name', 'type', 'status', 'created_on', 'updated_on',
            'supply_chain', 'data', 'validated_data', 'is_valid',
            'total_new_items', 'new_items_completed', 'new_items_failed',
            'total_updations', 'updations_completed', 'updations_failed')
        model = bulk_models.BulkUpload
    
    def validate(self, attrs):
        if self.instance:
            return attrs
        node = session.get_current_node()
        tenant = session.get_current_tenant()
        ttype = attrs.pop('type')
        tem = bulk_models.Template.get_tenant_template(template_type=ttype, tenant=tenant)
        try:
            wb = openpyxl.load_workbook(attrs['file'], data_only=True)
        except Exception as e:
            capture_exception(e)
            raise exceptions.BadRequest(_("Unrecognized file."))
        required_sheets = [tem.sheet_name, tem.processing_sheet_name, tem.constants_sheet_name]
        if not all(i in wb.sheetnames for i in required_sheets):
            raise exceptions.BadRequest(
                _("Invalid Excel. Please upload the originally downloaded file."))

        pws = wb[tem.processing_sheet_name]
        if pws[tem.tenant_id_cell].value != tenant.idencode:
            raise exceptions.BadRequest(_("Invalid Excel. Tenant Mismatch."))
        if pws[tem.node_id_cell].value != node.idencode:
            raise exceptions.BadRequest(_("Invalid Excel. Node Mismatch."))
        if pws[tem.sc_id_cell].value != attrs['supply_chain'].idencode:
            raise exceptions.BadRequest(_("Invalid Excel. SupplyChain Mismatch."))
        if pws[tem.type_cell].value != ttype:
            raise exceptions.BadRequest(_("Invalid Excel. Please check the file you uploaded."))
        attrs['template'] = tem
        attrs['node'] = node
        attrs['tenant'] = tenant
        return attrs

    def create(self, validated_data):
        validated_data['file_name'] = validated_data['file']._name
        bulk_upload = super(BulkUploadSerializer, self).create(validated_data)
        bulk_upload.validate()
        return bulk_upload
    
    def update(self, instance, validated_data):
        if instance.status not in [
                constants.BulkUploadStatuses.CREATED, constants.BulkUploadStatuses.VALIDATED]:
            raise exceptions.BadRequest(_("Bulk Upload already started. Cannot update"))

        # Only data can be updated
        [validated_data.pop(i) for i in validated_data.keys() - ('data',)]

        new = 0
        updation = 0
        for item in validated_data['data']:
            if 'id' in item:
                updation += 1
            else:
                new += 1
        validated_data['total_new_items'] = new
        validated_data['total_updations'] = updation
        bulk_upload = super(BulkUploadSerializer, self).update(instance, validated_data)
        bulk_upload.validate()
        return bulk_upload


class BulkUploadListSerializer(serializers.ModelSerializer):
    """
    Serializer to validate a template
    """
    id = fields.IdencodeField(read_only=True)
    type = serializers.IntegerField(read_only=True, source='template__type')

    class Meta:
        fields = (
            'id', 'file', 'file_name', 'type', 'status', 'created_on', 'updated_on',
            'total_new_items', 'new_items_completed', 'new_items_failed',
            'total_updations', 'updations_completed', 'updations_failed')
        model = bulk_models.BulkUpload



class TemplateFieldsSerializer(serializers.ModelSerializer):
    """
    Serializer to get all the fields with respect to the templates
    """

    label = serializers.CharField(source='field.name')
    placeholder = serializers.CharField(source='field.placeholder')
    required = serializers.BooleanField(source='field.required')
    type = serializers.IntegerField(source='field.type')

    class Meta:
        fields = ('label', 'placeholder', 'required', 'type', 'has_choices', 
            'column_pos', 'width', 'can_read', 'can_write', 'options' )
        model = bulk_models.TemplateField


class TemplateSerializer(serializers.ModelSerializer):
    """
    Serializer to create custom template and their respective template field 
    types.
    """
    id = fields.IdencodeField(read_only=True)
    tenant = fields.IdencodeField(
        related_model=tenant_models.Tenant, required=False, write_only=True)
    # titles = serializers.SerializerMethodField()
    column_data = serializers.ListField(required=False)
    template = fields.IdencodeField(
        related_model=bulk_models.Template, required=False, write_only=True)  
    
    # def get_titles(self, value):
    #     titles = []
    #     if value.file:
    #         file = value.file
    #         title_row = value.title_row
    #         wb = openpyxl.load_workbook(file, data_only=True)
    #         sheets = wb.sheetnames
    #         active_sheet = wb.active
    #         titles = []
    #         for cell in active_sheet[f'{title_row}']:
    #             if cell.value:
    #                 titles.append(cell.value)
    #     return titles

    def create(self, validated_data):
        validated_data['tenant'] = session.get_current_tenant()
        column_data = validated_data.pop('column_data')
        template = validated_data['template']
        for data in column_data:
            field = bulk_models.TemplateFieldType.objects.get(
                key=data['key'])
            bulk_models.TemplateField.objects.create(
                template=template, column_pos=data['column'], field=field)
        return template

    class Meta:
        fields = ('id', 'name', 'description', 'file_name', 'type', 'file', 
            'index_column', 'data_row', 'title_row', 'sheet_name', 'is_default', 
            'is_deleted', 'tenant', 'column_data', 'template')
        model = bulk_models.Template