
from openpyxl import styles

THIN_BORDER = styles.Border(
    left=styles.Side(style=styles.borders.BORDER_THIN),
    right=styles.Side(style=styles.borders.BORDER_THIN),
    top=styles.Side(style=styles.borders.BORDER_THIN),
    bottom=styles.Side(style=styles.borders.BORDER_THIN)
)
THICK_BORDER = styles.Border(
    left=styles.Side(style=styles.borders.BORDER_THICK),
    right=styles.Side(style=styles.borders.BORDER_THICK),
    top=styles.Side(style=styles.borders.BORDER_THICK),
    bottom=styles.Side(style=styles.borders.BORDER_THICK)
)
LEFT_THICK_BORDER = styles.Border(
    left=styles.Side(style=styles.borders.BORDER_THICK),
    right=styles.Side(style=styles.borders.BORDER_THIN),
    top=styles.Side(style=styles.borders.BORDER_THIN),
    bottom=styles.Side(style=styles.borders.BORDER_THIN)
)
RIGHT_THICK_BORDER = styles.Border(
    left=styles.Side(style=styles.borders.BORDER_THIN),
    right=styles.Side(style=styles.borders.BORDER_THICK),
    top=styles.Side(style=styles.borders.BORDER_THIN),
    bottom=styles.Side(style=styles.borders.BORDER_THIN)
)
TOP_THICK_BORDER = styles.Border(
    left=styles.Side(style=styles.borders.BORDER_THIN),
    right=styles.Side(style=styles.borders.BORDER_THIN),
    top=styles.Side(style=styles.borders.BORDER_THICK),
    bottom=styles.Side(style=styles.borders.BORDER_THIN)
)
BOTTOM_THICK_BORDER = styles.Border(
    left=styles.Side(style=styles.borders.BORDER_THIN),
    right=styles.Side(style=styles.borders.BORDER_THIN),
    top=styles.Side(style=styles.borders.BORDER_THIN),
    bottom=styles.Side(style=styles.borders.BORDER_THICK)
)
BOTTOM_LEFT_THICK_BORDER = styles.Border(
    left=styles.Side(style=styles.borders.BORDER_THICK),
    right=styles.Side(style=styles.borders.BORDER_THIN),
    top=styles.Side(style=styles.borders.BORDER_THIN),
    bottom=styles.Side(style=styles.borders.BORDER_THICK)
)
BOTTOM_RIGHT_THICK_BORDER = styles.Border(
    left=styles.Side(style=styles.borders.BORDER_THIN),
    right=styles.Side(style=styles.borders.BORDER_THICK),
    top=styles.Side(style=styles.borders.BORDER_THIN),
    bottom=styles.Side(style=styles.borders.BORDER_THICK)
)
PRIMARY_COLOUR = styles.Color(rgb="0F48AA")
PRIMARY_WHITE_COLOUR = styles.Color(rgb="FFFFFF")
DARK_COLOUR = styles.Color(rgb="CFD8E6")
LIGHT_COLOUR = styles.Color(rgb="E8ECF3")
TITLE_CELL = styles.PatternFill(fgColor=PRIMARY_COLOUR, fill_type="solid")
DARK_CELL = styles.PatternFill(fgColor=DARK_COLOUR, fill_type="solid")
LIGHT_CELL = styles.PatternFill(fgColor=LIGHT_COLOUR, fill_type="solid")

DATA_FONT = styles.Font(size="14", bold=False)
TITLE_FONT = styles.Font(size="18", color=PRIMARY_WHITE_COLOUR)
CENTER_ALIGN = styles.Alignment(horizontal='center', vertical='center')
LOCKED_CELL = styles.Protection(locked=True)
UNLOCKED_CELL = styles.Protection(locked=False)
