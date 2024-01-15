"""Models of the app Bulk Templates."""
import openpyxl
import sentry_sdk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, rows_from_range, cols_from_range
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import numbers as number_formats

from django.db import models
from django.db.models.functions import Length
from django.core.validators import MaxValueValidator
from django.utils.translation import gettext_lazy as _

from base.models import AbstractBaseModel
from base import session
from base import exceptions
from utilities import functions as util_functions
from common.library import _get_file_path
from v1.nodes import constants as node_constants
from v1.transactions import constants as transaction_constants
from v1.bulk_templates import constants as temp_consts
from v1.bulk_templates import notifications as bulk_notification

from . import excel_styles as styles
from . import tasks


class Template(AbstractBaseModel):
    """
    Model to save basic details of custom template of Tenants.
    Tenants can create their own bulk transaction excel templates
    to import transactions or connections.

    Attributes:
        tenant(obj)      : company objects that create template.
        name(str)       : name of template.
        title_row(str)  : row number where the title of the
                          template starts.
        data_row(str)   : row number where the data of the
                          template starts
        file(file)      : Uploaded file.
        type(int)       : type of template ie, what type of datas
                          are stored in this template. eg, template is
                          for txn or connection.
        default(bool)   : Is the template is the default template or not.
        is_deleted(bool): Is the template is deleted or not.

    Inherited Attribs:
        username(char): Username for the user for login.
        email(email): Email of the user.
        password(char): Password of the user.
        first_name(char): First name of the user.
        last_name(char): Last name of the user.
        date_joined(date): User added date.
    """
    tenant = models.ForeignKey(
        'tenants.Tenant', on_delete=models.CASCADE,
        related_name='excel_templates', verbose_name=_('Tenant'), null=True,
        blank=True)
    name = models.CharField(
        max_length=100, default='', blank=True, null=True,
        verbose_name=_('Template Name'))
    file_name = models.CharField(
        max_length=100, default='', blank=True, null=True,
        verbose_name=_('Template File Name'))
    description = models.CharField(
        max_length=1000, default='', null=True, blank=True,
        verbose_name=_('Excel Description'))
    sheet_name = models.CharField(
        max_length=100, default='Data', verbose_name=_('Sheet Name'))
    title_row = models.IntegerField(
        null=True, blank=True, default=4,
        verbose_name=_('Title Row Starting'))
    data_row = models.IntegerField(
        null=True, blank=True, default=5,
        verbose_name=_('Data Row Starting'))
    index_column = models.CharField(max_length=5, default='B')
    file = models.FileField(
        upload_to=_get_file_path, blank=True, null=True,
        verbose_name=_('Template'))
    type = models.IntegerField(
        choices=temp_consts.TemplateType.choices,
        default=temp_consts.TemplateType.TRANSACTION,
        verbose_name=_('Template Type'))
    is_default = models.BooleanField(
        default=False, verbose_name=_('Is Default Template'))
    is_deleted = models.BooleanField(
        default=False, verbose_name=_('Is Template Deleted'))

    constants_sheet_name = "Constants"
    processing_sheet_name = "Processor"
    data_row_height = 30
    desc_row_height = 50
    name_row = 2
    description_row = 3
    tenant_id_cell = 'A1'
    node_id_cell = 'B1'
    sc_id_cell = 'C1'
    extras_start_col = 'AA'
    type_cell = 'D1'
    item_count = 1000
    _backend_cur = 1
    _last_row = 0

    def __init__(self, *args, **kwargs):
        super(Template, self).__init__(*args, **kwargs)
        self._backend_cur = 1
        self.extras = {}
        self.generated_data_set = {}

    def __str__(self):
        """Object name in django admin."""
        return f"{self.name} | {self.id}"

    def get_template_config_class(self):
        return temp_consts.TemplateType(self.type).function()

    @classmethod
    def get_tenant_template(cls, template_type, tenant=None, id=None):
        tenant = tenant or session.get_current_tenant()        
        tenant_templates = tenant.excel_templates.filter(type=template_type)
        if id:
            tenant_templates = tenant_templates.filter(
                id=util_functions.decode(id))
        tem = tenant_templates.first() or \
            cls.objects.get(is_default=True, tenant=None, type=template_type)
        return tem

    def get_file_name(self):
        if not self.file_name:
            return f"{'_'.join(self.name.lower().split())}.xlsx"
        if self.file_name.endswith('.xlsx'):
            return self.file_name
        return f"{self.file_name.split('.')[0]}.xlsx"

    def get_n_data_columns(self, n=2):
        cols = (get_column_letter(self._backend_cur + i) for i in range(n))
        self._backend_cur += n
        return cols

    def index_cells(self):
        index_range = f"{self.index_column}{self.data_row}:{self.index_column}{self._last_row}"
        return list(cols_from_range(index_range))[0]

    def title_cells(self):
        title_range = f"{self.first_col}{self.title_row}:{self.last_col}{self.title_row}"
        return list(rows_from_range(title_range))[0]

    def data_cells(self):
        title_range = f"{get_column_letter(column_index_from_string(self.first_col)+1)}{self.data_row}:{self.data_br}"
        return rows_from_range(title_range)

    def field_col_cells(self, field):
        cell_range = f"{field.column_pos}{self.data_row}:{field.column_pos}{self._last_row}"
        return list(cols_from_range(cell_range))[0]

    def set_edges(self):
        self.first_col = self.index_column or self.template_fields.annotate(
            col_char_len=Length('column_pos')).order_by(
            'col_char_len', 'column_pos').first().column_pos
        self.last_col = self.template_fields.annotate(
            col_char_len=Length('column_pos')).order_by(
            'col_char_len', 'column_pos').last().column_pos
        self._last_row = self.data_row + self.item_count - 1
        self.data_tl = f"{self.first_col}{self.data_row}"  # Data top left cell
        self.data_tr = f"{self.last_col}{self.data_row}"  # Data top right cell
        self.data_bl = f"{self.first_col}{self._last_row}"  # Data bottom left cell
        self.data_br = f"{self.last_col}{self._last_row}"  # Data bottom right cell

    def add_title(self):
        self.ws[f'{self.first_col}{self.name_row}'] = self.name.upper()
        self.ws[f'{self.first_col}{self.description_row}'] = self.description
        self.ws.merge_cells(f'{self.first_col}{self.name_row}:{self.last_col}{self.name_row}')
        self.ws.merge_cells(f'{self.first_col}{self.description_row}:{self.last_col}{self.description_row}')

        for tem_field in self.template_fields.select_related('field').all():
            cell = f"{tem_field.column_pos}{self.title_row}"
            self.ws[cell] = f"{tem_field.field.name}{'*' if tem_field.field.required else ''}"

    def set_colours(self):
        self.ws[f'{self.first_col}{self.name_row}'].fill = styles.TITLE_CELL
        alternating_colours = [styles.DARK_CELL, styles.LIGHT_CELL]
        data_range = f"{self.data_tl}:{self.data_br}"
        for i, row in enumerate(rows_from_range(data_range)):
            for cell in row:
                self.ws[cell].fill = alternating_colours[i % 2]
        for cell in self.title_cells() + self.index_cells():
            self.ws[cell].fill = styles.TITLE_CELL
        return True

    def set_dimensions(self):
        corner_cell = f"{get_column_letter(column_index_from_string(self.index_column)+1)}{self.data_row}"
        self.ws.freeze_panes = self.ws[corner_cell]
        self.ws.row_dimensions[self.name_row].height = self.data_row_height
        self.ws.row_dimensions[self.description_row].height = self.desc_row_height
        self.ws.row_dimensions[self.title_row].height = self.data_row_height
        for r in range(self.data_row, self._last_row):
            self.ws.row_dimensions[r].height = self.data_row_height
        for tem_field in self.template_fields.select_related('field').all():
            self.ws.column_dimensions[tem_field.column_pos].width = \
                tem_field.field.required_width
        return True

    def add_borders(self):
        self.ws[f'{self.first_col}{self.name_row}'].border = styles.THICK_BORDER
        self.ws[f'{self.first_col}{self.description_row}'].border = styles.THICK_BORDER
        for c in self.title_cells():
            self.ws[c].border = styles.THICK_BORDER
        first_col_range = f"{self.data_tl}:{self.data_bl}"
        for c in list(cols_from_range(first_col_range))[0]:
            self.ws[c].border = styles.LEFT_THICK_BORDER
        last_col_range = f"{self.data_tr}:{self.data_br}"
        for c in list(cols_from_range(last_col_range))[0]:
            self.ws[c].border = styles.RIGHT_THICK_BORDER
        end_range = f"{self.data_bl}:{self.data_br}"
        for c in list(rows_from_range(end_range))[0]:
            self.ws[c].border = styles.BOTTOM_THICK_BORDER
        self.ws[f"{self.data_bl}"].border = styles.BOTTOM_LEFT_THICK_BORDER
        self.ws[f"{self.data_br}"].border = styles.BOTTOM_RIGHT_THICK_BORDER

        data_second_cell = f"{get_column_letter(column_index_from_string(self.first_col)+1)}{self.data_row}"
        data_second_last_cell = f"{get_column_letter(column_index_from_string(self.last_col)-1)}{self._last_row-1}"
        for row in rows_from_range(f"{data_second_cell}:{data_second_last_cell}"):
            for cell in row:
                self.ws[cell].border = styles.THIN_BORDER
        return True

    def set_fonts(self):
        self.ws.font = styles.DATA_FONT
        self.ws[f'{self.first_col}{self.name_row}'].font = styles.TITLE_FONT
        self.ws[f'{self.first_col}{self.description_row}'].font = styles.DATA_FONT
        self.ws[f'{self.first_col}{self.name_row}'].alignment = styles.CENTER_ALIGN
        self.ws[f'{self.first_col}{self.description_row}'].alignment = styles.CENTER_ALIGN
        for cell in self.index_cells() + self.title_cells():
            self.ws[cell].alignment = styles.CENTER_ALIGN
            self.ws[cell].font = styles.TITLE_FONT
        for row in self.data_cells():
            for cell in row:
                self.ws[cell].font = styles.DATA_FONT
        return True

    def protect_sheet(self):
        self.ws.protection.sheet = True
        self.wb[self.constants_sheet_name].protection.sheet = True
        self.wb[self.processing_sheet_name].protection.sheet = True
        data_range = f"{self.data_tl}:{self.data_br}"
        for row in cols_from_range(data_range):
            for cell in row:
                self.ws[cell].protection = styles.UNLOCKED_CELL
        for cell in self.index_cells():
            self.ws[cell].protection = styles.LOCKED_CELL
        for pk_field in self.template_fields.select_related('field').filter(can_write=False):
            pk_col_range = f"{pk_field.column_pos}{self.data_row}:{pk_field.column_pos}{self._last_row}"
            for cell in list(cols_from_range(pk_col_range))[0]:
                self.ws[cell].protection = styles.LOCKED_CELL
        return True

    def style_sheet(self):
        self.add_title()
        self.set_colours()
        self.set_dimensions()
        self.add_borders()
        self.set_fonts()
        return self.ws

    def add_index(self):
        for i in range(self.data_row, self.data_row + self.item_count):
            index_cell = f"{self.index_column}{i}"
            self.ws[index_cell] = str(i - self.data_row + 1)

    def setup_file(self):
        """ Returns the base file for this template."""
        if self.file:
            self.wb = openpyxl.load_workbook(self.file)
            self.ws = self.wb[self.sheet_name]
        else:
            self.wb = Workbook()
            self.wb.remove_sheet(self.wb.active)

            if self.sheet_name not in self.wb.sheetnames:
                self.ws = self.wb.create_sheet(self.sheet_name)
            self.style_sheet()

        if self.constants_sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(self.constants_sheet_name)
        self.wb[self.constants_sheet_name].sheet_state = 'hidden'

        if self.processing_sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(self.processing_sheet_name)
        self.wb[self.processing_sheet_name].sheet_state = 'hidden'
        self.protect_sheet()
        return self.wb

    def create_named_range(self, key, data):
        bws = self.wb[self.constants_sheet_name]
        c1, c2 = self.get_n_data_columns(2)
        enum_i = 0
        for enum_i, item in enumerate(data):
            bws[f'{c1}{enum_i+1}'] = item[0]
            bws[f'{c2}{enum_i+1}'] = item[1]
        formula = f'${c1}$1:${c1}${enum_i+1}'
        try:
            self.wb.create_named_range(key, bws, formula) #TODO: REWORK
        except:
            pass
        return f"{self.constants_sheet_name}!{c1}1:{c2}{enum_i+1}"

    def create_dropdown(self, tem_field):
        pws = self.wb[self.processing_sheet_name]
        data_range = self.create_named_range(
            tem_field.field.key,
            tem_field.field.get_generated_data(
                self._tenant, self._node, self._supply_chain, **self.extras)
        )

        for i in range(self.data_row, self.data_row + self.item_count):
            pws[f"{tem_field.column_pos}{i}"] = \
                f'=IFERROR(VLOOKUP({self.sheet_name}!{tem_field.column_pos}{i}, ' \
                f'{data_range}, 2, FALSE), "")'
        return True
    
    def add_validations(self):
        for tem_field in self.template_fields.select_related('field').all():
            first_cell = f'{tem_field.column_pos}{self.data_row}'
            if tem_field.field.data_generator:
                self.create_dropdown(tem_field)

            data_val = tem_field.field.get_excel_validation(first_cell)
            if data_val:
                self.ws.add_data_validation(data_val)
                validation_range = f'{tem_field.column_pos}{self.data_row}:{tem_field.column_pos}{self._last_row}'
                data_val.add(validation_range)
            tem_field.format_col_cells(self)

    def annotate_metadata(self):
        pws = self.wb[self.processing_sheet_name]
        pws[self.tenant_id_cell] = self._tenant.idencode
        pws[self.node_id_cell] = self._node.idencode
        pws[self.sc_id_cell] = self._supply_chain.idencode
        pws[self.type_cell] = self.type

        return True

    def add_data(self):
        TemplateConfigClass = self.get_template_config_class()
        dataset = TemplateConfigClass.data(self._tenant, self._node, self._supply_chain, **self.extras)
        for i, item in enumerate(dataset):
            row = self.data_row + i
            for tem_field in self.template_fields.select_related('field').filter(can_read=True):
                cell = f"{tem_field.column_pos}{row}"
                val = item.get(tem_field.field.key, "#ERROR")
                self.ws[cell] = str(val) if val or type(val) in [int, float, bool] else ""

    def get_template_file(self, supply_chain, tenant=None, node=None, **kwargs):
        """ Fetches the base file and populated headers, name description etc"""
        self._tenant = tenant or session.get_current_tenant()
        self._node = node or session.get_current_node()
        self._supply_chain = supply_chain
        self.extras = kwargs
        TemplateConfigClass = self.get_template_config_class()

        self.set_edges()

        self.setup_file()
        self.annotate_metadata()
        self.add_index()
        self.add_validations()
        self.add_data()
        TemplateConfigClass.customize_file(self)
        self.ws.protection.sheet = True
        return openpyxl.writer.excel.save_virtual_workbook(self.wb)
    
    def get_extras(self, wb):
        pws = wb[self.processing_sheet_name]
        i = 0
        d = {}
        while True:
            col = get_column_letter(column_index_from_string(self.extras_start_col) + i)
            key = pws[f'{col}1'].value
            value = pws[f'{col}2'].value
            if not key:
                break
            d[key] = value
            i += 1
        return d

    def get_generated_data(self, tenant, node, supply_chain, generator_type, field):
        """
        This function was moved here from TemplateFieldType.get_generated_data
        to optimize the code and avoid re-generation of data for the field in each row.
        """
        if not generator_type:
            return []
        if generator_type in self.generated_data_set:
            return self.generated_data_set[generator_type]
        self.generated_data_set[generator_type] = field.get_generated_data(
            tenant, node, supply_chain, **self.extras)
        return self.generated_data_set[generator_type]

    def validate_data(self, data, tenant, node, supply_chain):
        sheet_valid = True
        sheet_data = []
        error_count = 0
        for row in data:
            row_error_count = 0
            row_valid = True
            row_data = {}

            for tem_field in self.template_fields.select_related('field').filter():
                value = row.get(tem_field.field.key, None)
                cell_validation = tem_field.field.validate(
                    value,
                    self.get_generated_data(
                        tenant, node, supply_chain, tem_field.field.data_generator, tem_field.field)
                )
                row_error_count += not cell_validation['is_valid']
                row_data[tem_field.field.key] = cell_validation
                row_valid = row_valid and cell_validation['is_valid']
                row[tem_field.field.key] = cell_validation['value']
            row_data['force_create'] = {
                "value": row.get('force_create', False),
                "message": "",
                "is_valid": True,
                "hidden": True,
            }
            row_data['_status'] = {
                "value": temp_consts.BulkUploadStatuses.VALIDATED,
                "message": "",
                "is_valid": True,
                "hidden": True,
            }

            sheet_data.append({
                'data': row_data,
                'is_valid': row_valid,
                'message': '',
                'error_count': row_error_count
            })
            sheet_valid = sheet_valid and row_valid
            error_count += row_error_count

        TemplateConfigClass = self.get_template_config_class()
        sheet_data, valid = TemplateConfigClass.validate(
            sheet_data, tenant, node, supply_chain, **self.extras)

        return util_functions.serialize_promises(sheet_data), sheet_valid and valid

    def extract_data(self, excel_file):
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        self.extras = self.get_extras(wb)
        self.ws = wb[self.sheet_name]
        pws = wb[self.processing_sheet_name]
        data = []
        emtpy_rows = 0
        row_index = self.data_row
        while emtpy_rows <= 10:
            row_filled = False
            item = {}
            for tem_field in self.template_fields.select_related('field').all():
                cell = f"{tem_field.column_pos}{row_index}"
                if tem_field.field.data_generator:
                    cell_value = pws[cell].value
                else:
                    cell_value = self.ws[cell].value
                item[tem_field.field.key] = cell_value
                row_filled = row_filled or (
                        bool(cell_value) and
                        cell_value is not False and
                        tem_field.field.key not in self.extras.keys()
                )
            if row_filled:
                emtpy_rows = 0
                data.append(item)
            else:
                emtpy_rows += 1
            row_index += 1

        return data


class TemplateFieldType(AbstractBaseModel):
    """
    Model to save field details of template. field have
    different properties like name, type, description etc. also
    can set field is required or not.

    Attributes:
        template_type(int)  : type of template ie, what type of data
                              are stored in this field. eg, txn data
                              or connection.
        name(str)           : name of field.
        type(int)           : type of field. eg: char, int etc.
        description(str)    : description of field.
        required(bool)      : set field is required or not.
        key(char)           : key name for the field.
        is_dynamic(bool)    : value can be taken from frontend also.

    """
    required = models.BooleanField(
        default=False, verbose_name=_('Is Field Required'))
    name = models.CharField(max_length=500, verbose_name=_('Name Of Field'))
    description = models.CharField(
        max_length=1000, default='', null=True, blank=True,
        verbose_name=_('Description About Field'))
    template_type = models.IntegerField(
        choices=temp_consts.TemplateType.choices,
        default=temp_consts.TemplateType.TRANSACTION,
        verbose_name=_('Template Type'))
    type = models.IntegerField(
        choices=temp_consts.TemplateFieldTypes.choices,
        default=temp_consts.TemplateFieldTypes.STRING,
        verbose_name=_('Field Type'))
    width = models.PositiveIntegerField(
        validators=[MaxValueValidator(30)], default=5,
        verbose_name=_('Width of cell'))

    key = models.CharField(max_length=1000, verbose_name=_('Field Key Name'))
    data_generator = models.IntegerField(
        choices=temp_consts.DataGenerator.choices,
        default=None, null=True, blank=True,
        verbose_name=_('Data Generator'))
    placeholder = models.CharField(max_length=1000, null=True, blank=True)
    is_dynamic = models.BooleanField(default=False)

    class Meta:
        """Meta Info."""
        unique_together = ('name', 'key',)

    def __str__(self):
        """Object name in django admin."""
        return f"{self.name} - {self.get_type_display()} - {self.idencode}"

    def get_generated_data(self, tenant, node, supply_chain, **kwargs):
        """
        Get generated data as list of tuples if a data generator is selected.
        """
        if not self.data_generator:
            raise ValueError(f"No data generator configured for Field Type {self.name}")
        return temp_consts.DataGenerator(self.data_generator).function(
            tenant, node, supply_chain, **kwargs)

    def validate(self, value, generated_data):
        data = temp_consts.TemplateFieldTypes(self.type).function(value, self.required)
        data['display_text'] = self.get_display_text(
            data['value'], generated_data)
        data['label'] = self.name
        data['key'] = self.key
        data['place_holder'] = self.placeholder
        data['width'] = self.width
        data['type'] = self.type
        data['is_required'] = self.required
        data['options'] = dict(generated_data)
        data['is_dynamic'] = self.is_dynamic
        return data

    @property
    def required_width(self):
        """
        Numbers are found with minimal testing.
        Possibility of mismatch for different type os cells exists.
        """
        return max(len(self.name)*1.7, self.width * 3)

    def get_display_text(self, value, generated_data):
        if not self.data_generator:
            return value
        options_dict = dict([i[::-1] for i in generated_data])
        return options_dict.get(value, value)

    def get_excel_validation(self, first_cell='A1'):
        if self.data_generator:
            return DataValidation(
                type="list", formula1=f'={self.key}', allow_blank=True,
                errorTitle=f"Invalid {self.name}",
                error=f"Select {self.name} from dropdown.")
        if self.type == temp_consts.TemplateFieldTypes.DATE:
            return DataValidation(
                type="date", allow_blank=True,
                errorTitle=f"Invalid {self.name}", error=f"Input a valid date.")
        if self.type in [
                temp_consts.TemplateFieldTypes.FLOAT,
                temp_consts.TemplateFieldTypes.INTEGER]:
            return DataValidation(
                type="decimal", allow_blank=True,
                errorTitle=f"Invalid {self.name}",
                error=f"Input a valid decimal value.")
        if self.type in [
                temp_consts.TemplateFieldTypes.POSITIVE_INTEGER,
                temp_consts.TemplateFieldTypes.POSITIVE_FLOAT]:
            return DataValidation(
                type="decimal", formula1=0, allow_blank=True,
                operator="greaterThan",
                errorTitle=f"Invalid {self.name}",
                error=f"Input a valid positive decimal value.")
        if self.type == temp_consts.TemplateFieldTypes.EMAIL:
            email_formula = f'=AND(ISERROR(FIND(" ",{first_cell})),LEN({first_cell})-LEN(' \
                            f'SUBSTITUTE({first_cell},"@",""))=1,IFERROR(SEARCH("@",{first_cell})' \
                            f'<SEARCH(".",{first_cell},SEARCH("@",{first_cell})),0),ISERROR(FIND(' \
                            f'",",{first_cell})),NOT(IFERROR(SEARCH(".",{first_cell},SEARCH("@",{first_cell}))' \
                            f'-SEARCH("@",{first_cell}),0)=1),LEFT({first_cell},1)<>".",RIGHT({first_cell},1)<>".")'
            return DataValidation(
                type="custom", formula1=email_formula, allow_blank=True,
                errorTitle=f"Invalid {self.name}",
                error=f"Input a valid email address.")

        return None

    def format_cells(self, cells):
        cell_format = None
        if self.type in [
                temp_consts.TemplateFieldTypes.INTEGER,
                temp_consts.TemplateFieldTypes.POSITIVE_INTEGER]:
            cell_format = number_formats.FORMAT_NUMBER
        if self.type in [
                temp_consts.TemplateFieldTypes.FLOAT,
                temp_consts.TemplateFieldTypes.POSITIVE_FLOAT]:
            cell_format = "0.000"
        if self.type == temp_consts.TemplateFieldTypes.DATE:
            cell_format = "dd-mm-yyyy"
        if self.type == temp_consts.TemplateFieldTypes.PHONE:
            cell_format = number_formats.FORMAT_TEXT
        if cell_format:
            for cell in cells:
                cell.number_format = cell_format
        return True


class TemplateField(AbstractBaseModel):
    """
    model to store the column position of each template field.
    tenant can select the column and field with their own choice when
    creating a template.

     Attributes:
        template(obj)       : template object.
        field(obj)          : template type field object.
        column_pos(str)     : column position of a field that the user
                              select when creating custom template.

    """
    template = models.ForeignKey(
        Template, on_delete=models.CASCADE,
        related_name='template_fields',
        null=True, blank=True, default=None,
        verbose_name=_('Template'))
    field = models.ForeignKey(
        TemplateFieldType, on_delete=models.CASCADE,
        related_name='template_field',
        null=True, blank=True, default=None,
        verbose_name=_('Field'))
    column_pos = models.CharField(
        max_length=4, null=True, blank=True,
        default='A', verbose_name=_('Column Position'))
    width = models.PositiveIntegerField(
        validators=[MaxValueValidator(12)], default=5,
        verbose_name=_('Width of cell'))
    can_read = models.BooleanField(default=True, verbose_name=_('Can Read'))
    can_write = models.BooleanField(default=True, verbose_name=_('Can Write'))

    def __str__(self):
        """Object name in django admin."""
        return f"{self.field.name} - {self.column_pos} - {self.idencode}"

    def format_col_cells(self, template):
        cells = template.field_col_cells(self)
        self.field.format_cells([template.ws[cell] for cell in cells])
        return True
    
    def has_choices(self):
        """Returns true if the field has choices"""

        return bool(self.field.data_generator)
    
    @property
    def options(self):
        """Return options for the respective fields"""
        
        if not self.field.data_generator:
            return []
        tenant = session.get_current_tenant()
        node = session.get_current_node()
        supply_chain = session.get_current_user().get_default_sc()
        options = self.template.get_generated_data(
            tenant, node, supply_chain, self.field.data_generator, self.field)
        return options


class BulkUpload(AbstractBaseModel):
    """
    Model to store the bulk upload files and also track the progress of adding
    connections.

    Attributes:
        node(obj)               : Node that is doing the operation
        file(file)              : File that is uploaded
        data(json)              : Data that is uploaded
        errors(json)            : Errors in the data
        actors_to_add(int)     : Number of new actors in the excel
        actors_added(int)      : Actors successfully added
        actors_to_update(int)  : Actors to be updated
        actors_updated(int)    : Actors successfully updated
        transactions_to_add(int): Transactions to be added
        transactions_added(int) : Transactions successfully added
        used(bool)              : Whether the file is used or not.


    """
    tenant = models.ForeignKey(
        'tenants.Tenant', on_delete=models.CASCADE,
        default=None, null=True, blank=True,
        related_name='bulk_uploads', verbose_name=_("Tenant"))
    node = models.ForeignKey(
        'nodes.Node', on_delete=models.CASCADE,
        related_name='bulk_uploads', verbose_name=_('Node'),
        null=True, blank=True)
    supply_chain = models.ForeignKey(
        'supply_chains.SupplyChain', on_delete=models.CASCADE,
        default=None, null=True, blank=True,
        related_name='bulk_uploads', verbose_name=_("Supply Chain"))
    template = models.ForeignKey(
        Template, on_delete=models.CASCADE,
        related_name='bulk_uploads',
        null=True, blank=True, default=None,
        verbose_name=_('Template'))
    file = models.FileField(
        upload_to=_get_file_path, blank=True, null=True,
        verbose_name=_('Excel File'))
    file_name = models.CharField(
        max_length=100, verbose_name=_('File Name'))
    validated_data = models.JSONField(
        null=True, blank=True, default=list, verbose_name=_('Validated Data'))
    data = models.JSONField(
        null=True, blank=True, default=list, verbose_name=_('Data'))
    errors = models.JSONField(
        null=True, blank=True, default=list, verbose_name=_('Error Data'))
    status = models.IntegerField(
        choices=temp_consts.BulkUploadStatuses.choices,
        default=temp_consts.BulkUploadStatuses.CREATED,
        verbose_name=_('Status'),)
    total_new_items = models.IntegerField(
        default=0, verbose_name=_('New Items'))
    new_items_completed = models.IntegerField(
        default=0, verbose_name=_('New Items Completed'))
    new_items_failed = models.IntegerField(
        default=0, verbose_name=_('New Items Failed'))
    total_updations = models.IntegerField(
        default=0, verbose_name=_('Total Updations'))
    updations_completed = models.IntegerField(
        default=0, verbose_name=_('Updations Completed'))
    updations_failed = models.IntegerField(
        default=0, verbose_name=_('Updations Failed'))
    is_valid = models.BooleanField(
        default=False, verbose_name=_('Is file data valid'))
    is_used = models.BooleanField(
        default=False, verbose_name=_('Is Data Import Completed'))

    def __str__(self):
        """Object name in django admin."""
        return f"Bulk operation - {self.node.name} - {self.idencode}"

    def validate(self, **kwargs):
        """
        Validates the bulk upload file against the template.
        """
        _data = []
        if self.status == temp_consts.BulkUploadStatuses.CREATED:
            _data = self.template.extract_data(self.file)
        elif self.status == temp_consts.BulkUploadStatuses.VALIDATED:
            _data = self.data
        if not _data:
            raise exceptions.BadRequest(_("Excel is empty."))
        self.validated_data, self.is_valid = \
            self.template.validate_data(
                _data, self.tenant, self.node, self.supply_chain, **kwargs)

        self.data = [
            {k: v['value'] for k, v in i['data'].items()}
            for i in self.validated_data
        ]

        self.status = temp_consts.BulkUploadStatuses.VALIDATED
        self.save()
        return self.is_valid

    def start_upload(self,extra_data=None):
        """
        Function to start the celery task to upload data in bulk.
        """
        if not self.is_valid:
            raise exceptions.BadRequest(_("Template still has validation errors."))
        self.total_updations = 0
        self.total_new_items = 0
        if self.status in [temp_consts.BulkUploadStatuses.INITIATED,
                           temp_consts.BulkUploadStatuses.IN_PROGRESS]:
            raise exceptions.Conflict(_("Bulk upload is already in progress"))

        for item in self.data:
            if 'id' in item and item['id'] and item['id'] != "None":
                self.total_updations += 1
            else:
                self.total_new_items += 1
        for item in self.validated_data:
            item['data']['_status'] = {
                "value": temp_consts.BulkUploadStatuses.INITIATED,
                "message": "",
                "is_valid": True,
                "hidden": True,
            }
        self.status = temp_consts.BulkUploadStatuses.INITIATED
        self.errors = []
        self.save()
        tasks.upload.delay(upload_id=self.id,extra_data=extra_data)
        return True

    def notify(self):
        """
        Notify bulk-upload created user when the upload is completed or failed.
        """
        notify_statuses = [
            temp_consts.BulkUploadStatuses.COMPLETED, 
            temp_consts.BulkUploadStatuses.FAILED, 
            ]
        if self.status in notify_statuses:
            notification_manager = bulk_notification.BulkUploadNotificationManager(
                user=self.creator, action_object=self, token=None)
            notification_manager.send_notification()
        return True
